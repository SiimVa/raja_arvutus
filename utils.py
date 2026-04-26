import math
import requests
from datetime import datetime, timedelta, time, date
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
from geopy.distance import geodesic
import mgrs as mgrs_lib
import folium
from astral import LocationInfo
from astral.sun import sun, sunrise as astral_sunrise, sunset as astral_sunset
import plotly.express as px
import plotly.graph_objects as go


# ======================================================
# ABI- JA PARSIFUNKTSIOONID
# ======================================================

def parse_datetime(dt_str: str) -> datetime:
    return datetime.strptime(dt_str, "%Y-%m-%d %H:%M")


def parse_date(d_str: str) -> date:
    return datetime.strptime(d_str, "%Y-%m-%d").date()


def parse_clock(t_str: str) -> time:
    return datetime.strptime(t_str, "%H:%M").time()


def minutes_between(dt1: datetime, dt2: datetime) -> float:
    return (dt2 - dt1).total_seconds() / 60.0


def mgrs_to_latlon(mgrs_code: str) -> Tuple[float, float]:
    converter = mgrs_lib.MGRS()
    lat, lon = converter.toLatLon(mgrs_code)
    return float(lat), float(lon)


def normalize_control_point_durations(control_points: pd.DataFrame) -> pd.DataFrame:
    cp = control_points.copy()
    detailed_cols = {"kestvus_ettevalmistus_min", "kestvus_uleanne_min", "kestvus_tagasiside_min"}

    if detailed_cols.issubset(cp.columns):
        cp["kestvus_min"] = (
            cp["kestvus_ettevalmistus_min"].fillna(0).astype(float)
            + cp["kestvus_uleanne_min"].fillna(0).astype(float)
            + cp["kestvus_tagasiside_min"].fillna(0).astype(float)
        )
    elif "kestvus_min" in cp.columns:
        cp["kestvus_ettevalmistus_min"] = 0
        cp["kestvus_uleanne_min"] = cp["kestvus_min"].astype(float)
        cp["kestvus_tagasiside_min"] = 0
    else:
        raise ValueError("Kontrollpunktide tabel peab sisaldama kas kestvus_min või kestvus_ettevalmistus_min;kestvus_uleanne_min;kestvus_tagasiside_min")
    return cp


def straight_distance_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    return geodesic((lat1, lon1), (lat2, lon2)).meters


# ======================================================
# PÄIKESETÕUS / PÄIKESELOOJANG
# ======================================================

def enrich_control_points_with_coordinates(control_points: pd.DataFrame) -> pd.DataFrame:
    cp = control_points.copy().sort_values("jarjekord").reset_index(drop=True)
    lats = []
    lons = []
    for code in cp["mgrs"]:
        lat, lon = mgrs_to_latlon(code)
        lats.append(lat)
        lons.append(lon)
    cp["lat"] = lats
    cp["lon"] = lons
    return cp


def get_reference_location(control_points: pd.DataFrame) -> Tuple[float, float]:
    lat = float(control_points["lat"].mean())
    lon = float(control_points["lon"].mean())
    return lat, lon


def compute_sun_times(control_points: pd.DataFrame, race_config: dict) -> dict:
    cp = enrich_control_points_with_coordinates(control_points)
    ref_lat, ref_lon = get_reference_location(cp)
    race_date = parse_date(race_config["voistluse_kuupaev"])
    timezone_name = race_config.get("timezone", "Europe/Tallinn")

    location = LocationInfo(name="RaceArea", region="", timezone=timezone_name, latitude=ref_lat, longitude=ref_lon)
    s = sun(location.observer, date=race_date, tzinfo=timezone_name)

    race_config_out = dict(race_config)
    race_config_out["paeva_algus"] = s["sunrise"].strftime("%H:%M")
    race_config_out["pimeduse_algus"] = s["sunset"].strftime("%H:%M")
    race_config_out["sunrise_full"] = s["sunrise"]
    race_config_out["sunset_full"] = s["sunset"]
    race_config_out["reference_lat"] = ref_lat
    race_config_out["reference_lon"] = ref_lon
    return race_config_out


def get_sun_period_datetimes(race_config: dict) -> tuple[datetime, datetime]:
    if "sunrise_full" in race_config and "sunset_full" in race_config:
        sunrise = race_config["sunrise_full"]
        sunset = race_config["sunset_full"]
    else:
        race_date = parse_date(race_config["voistluse_kuupaev"])
        sunrise = datetime.combine(race_date, parse_clock(race_config["paeva_algus"]))
        sunset = datetime.combine(race_date, parse_clock(race_config["pimeduse_algus"]))
    return sunrise, sunset


def get_sun_period_datetimes_for_date(race_config: dict, target_date: date) -> tuple[datetime, datetime]:
    timezone_name = race_config.get("timezone", "Europe/Tallinn")
    if "reference_lat" in race_config and "reference_lon" in race_config:
        location = LocationInfo(
            name="RaceArea",
            region="",
            timezone=timezone_name,
            latitude=float(race_config["reference_lat"]),
            longitude=float(race_config["reference_lon"]),
        )
        sunrise_dt = astral_sunrise(location.observer, date=target_date, tzinfo=timezone_name)
        sunset_dt = astral_sunset(location.observer, date=target_date, tzinfo=timezone_name)
        return sunrise_dt.replace(tzinfo=None), sunset_dt.replace(tzinfo=None)

    sunrise_time = parse_clock(race_config["paeva_algus"])
    sunset_time = parse_clock(race_config["pimeduse_algus"])
    return datetime.combine(target_date, sunrise_time), datetime.combine(target_date, sunset_time)


# ======================================================
# DISTANTSIDE LEIDMINE
# ======================================================

def road_distance_m_osrm(lat1: float, lon1: float, lat2: float, lon2: float, timeout: int = 15) -> Optional[Tuple[float, List[Tuple[float, float]]]]:
    url = (
        "https://router.project-osrm.org/route/v1/driving/"
        f"{lon1},{lat1};{lon2},{lat2}?overview=full&geometries=geojson"
    )
    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        data = response.json()
        routes = data.get("routes", [])
        if routes:
            route = routes[0]
            distance = float(route["distance"])
            geometry = route.get("geometry", {})
            if isinstance(geometry, dict) and "coordinates" in geometry:
                coords = [(lat, lon) for lon, lat in geometry["coordinates"]]
                return distance, coords
            elif isinstance(geometry, str):
                # OSRM may still return an encoded polyline; fallback to direct line
                return distance, [(lat1, lon1), (lat2, lon2)]
            else:
                return distance, [(lat1, lon1), (lat2, lon2)]
    except Exception:
        return None
    return None


def calculate_segment_distances(control_points: pd.DataFrame, segments: pd.DataFrame) -> pd.DataFrame:
    cp_lookup = control_points.set_index("kp_id").to_dict("index")
    seg = segments.copy().sort_values("segment_id").reset_index(drop=True)

    straight_distances = []
    road_distances = []
    used_distances = []
    notes = []
    route_coords = []

    for _, row in seg.iterrows():
        start_cp = cp_lookup[row["algus_kp_id"]]
        end_cp = cp_lookup[row["lopp_kp_id"]]

        lat1, lon1 = start_cp["lat"], start_cp["lon"]
        lat2, lon2 = end_cp["lat"], end_cp["lon"]

        straight_m = straight_distance_m(lat1, lon1, lat2, lon2)
        road_result = None
        used_m = None
        note = ""
        coords = None

        if row["liikumisviis"] == "tee":
            road_result = road_distance_m_osrm(lat1, lon1, lat2, lon2)
            if road_result is None:
                road_m = straight_m * 1.2
                note = "OSRM ebaõnnestus, kasutati varuplaani: linnulend * 1.2"
                coords = [(lat1, lon1), (lat2, lon2)]
            else:
                road_m, coords = road_result
                note = "Tee-distants OSRM-ist"
            used_m = road_m
        else:
            used_m = straight_m * 1.5
            note = "Varjatud liikumine: linnulend * 1.5"
            coords = [(lat1, lon1), (lat2, lon2)]

        straight_distances.append(straight_m)
        road_distances.append(road_m if road_result else None)
        used_distances.append(used_m)
        notes.append(note)
        route_coords.append(coords)

    seg["sirge_kaugus_m"] = straight_distances
    seg["tee_kaugus_m"] = road_distances
    seg["kasutatav_kaugus_m"] = used_distances
    seg["distance_note"] = notes
    seg["route_coords"] = route_coords
    return seg


# ======================================================
# KIIRUSTE RAKENDAMINE
# ======================================================

def apply_speeds(segments: pd.DataFrame, default_speeds: dict, overrides: Dict[int, Dict[str, float]]) -> pd.DataFrame:
    seg = segments.copy()
    valge_list = []
    pime_list = []

    for _, row in seg.iterrows():
        seg_id = int(row["segment_id"])
        mode = row["liikumisviis"]

        default_valge = float(default_speeds[mode]["valge"])
        default_pime = float(default_speeds[mode]["pime"])
        final_valge = float(overrides.get(seg_id, {}).get("valge", default_valge))
        final_pime = float(overrides.get(seg_id, {}).get("pime", default_pime))

        if final_valge <= 0 or final_pime <= 0:
            raise ValueError(f"Lõigu {seg_id} kiirused peavad olema positiivsed.")

        valge_list.append(final_valge)
        pime_list.append(final_pime)

    seg["kiirus_valges_kmh"] = valge_list
    seg["kiirus_pimedas_kmh"] = pime_list
    return seg


# ======================================================
# STARDIAJAD
# ======================================================

def generate_team_start_times(race_config: dict) -> pd.DataFrame:
    first_start = parse_datetime(race_config["esimese_voistkonna_start"])
    interval = int(race_config["stardi_intervall_min"])
    earlier = int(race_config["nulltiimi_earlier_min"])
    team_count = int(race_config["voistkondade_arv"])

    rows = [{"team_id": 0, "start_time": first_start - timedelta(minutes=earlier)}]
    for i in range(1, team_count + 1):
        rows.append({"team_id": i, "start_time": first_start + timedelta(minutes=(i - 1) * interval)})
    return pd.DataFrame(rows)


# ======================================================
# VALGE / PIME LOOGIKA
# ======================================================

def is_light(dt: datetime, race_config: dict) -> bool:
    paeva_algus = parse_clock(race_config["paeva_algus"])
    pimeduse_algus = parse_clock(race_config["pimeduse_algus"])
    t = dt.time()
    return paeva_algus <= t < pimeduse_algus


def get_next_light_boundary(dt: datetime, race_config: dict) -> datetime:
    paeva_algus = parse_clock(race_config["paeva_algus"])
    pimeduse_algus = parse_clock(race_config["pimeduse_algus"])

    today = dt.date()
    today_day_start = datetime.combine(today, paeva_algus)
    today_dark_start = datetime.combine(today, pimeduse_algus)
    tomorrow_day_start = datetime.combine(today + timedelta(days=1), paeva_algus)

    if is_light(dt, race_config):
        return today_dark_start
    else:
        if dt < today_day_start:
            return today_day_start
        return tomorrow_day_start


def classify_interval(start_dt: datetime, end_dt: datetime, race_config: dict) -> str:
    start_light = is_light(start_dt, race_config)
    end_minus = end_dt - timedelta(seconds=1)
    end_light = is_light(end_minus, race_config)
    boundary = get_next_light_boundary(start_dt, race_config)

    if end_dt <= boundary and start_light == end_light:
        return "valge" if start_light else "pime"
    return "sega"


def calculate_segment_end_time(segment_distance_m: float, start_datetime: datetime, speed_white_kmh: float, speed_dark_kmh: float, race_config: dict) -> datetime:
    remaining_distance = float(segment_distance_m)
    current_time = start_datetime

    while remaining_distance > 0.01:
        current_is_light = is_light(current_time, race_config)
        next_boundary = get_next_light_boundary(current_time, race_config)

        speed_kmh = speed_white_kmh if current_is_light else speed_dark_kmh
        speed_m_per_min = speed_kmh * 1000.0 / 60.0
        available_minutes = max(minutes_between(current_time, next_boundary), 0.0)
        possible_distance = speed_m_per_min * available_minutes

        if possible_distance >= remaining_distance or available_minutes == 0:
            needed_minutes = remaining_distance / speed_m_per_min
            current_time = current_time + timedelta(minutes=needed_minutes)
            remaining_distance = 0.0
        else:
            remaining_distance -= possible_distance
            current_time = next_boundary

    return current_time


def calculate_fixed_segment_end_time(segment_distance_m: float, start_datetime: datetime, speed_kmh: float) -> datetime:
    speed_m_per_min = float(speed_kmh) * 1000.0 / 60.0
    needed_minutes = float(segment_distance_m) / speed_m_per_min
    return start_datetime + timedelta(minutes=needed_minutes)


def calculate_segment_minutes(segment_distance_m: float, speed_kmh: float) -> float:
    speed_m_per_min = float(speed_kmh) * 1000.0 / 60.0
    return float(segment_distance_m) / speed_m_per_min


def round_up_minutes_to_next_five(minutes: float) -> float:
    return float(math.ceil(float(minutes) / 5.0) * 5)


def determine_segment_condition(start_datetime: datetime, race_config: dict, team_count: int, interval_minutes: int) -> str:
    conditions = set()
    for team_index in [0, team_count - 1] if team_count > 1 else [0]:
        team_start = start_datetime + timedelta(minutes=team_index * interval_minutes)
        conditions.add("valge" if is_light(team_start, race_config) else "pime")
    return conditions.pop() if len(conditions) == 1 else "sega"


def round_up_to_next_five_minutes(dt: datetime) -> datetime:
    rounded = dt.replace(second=0, microsecond=0)
    if dt.second != 0 or dt.microsecond != 0 or dt.minute % 5 != 0:
        rounded += timedelta(minutes=(5 - dt.minute % 5))
    return rounded


# ======================================================
# SIMULATSIOON
# ======================================================

def simulate_team_route(team_id: int, team_start: datetime, control_points: pd.DataFrame, segments: pd.DataFrame, race_config: dict, start_duration_min: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    cp_lookup = control_points.set_index("kp_id").to_dict("index")
    seg_sorted = segments.sort_values("segment_id")

    segment_rows = []
    checkpoint_rows = []
    current_time = team_start + timedelta(minutes=start_duration_min)

    for _, seg in seg_sorted.iterrows():
        seg_id = int(seg["segment_id"])
        start_kp_id = int(seg["algus_kp_id"])
        end_kp_id = int(seg["lopp_kp_id"])
        dist_m = float(seg["kasutatav_kaugus_m"])
        exact_segment_minutes = float(seg["exact_segment_minutes"])
        rounded_segment_minutes = float(seg["rounded_segment_minutes"])

        seg_start = current_time
        seg_end_exact = seg_start + timedelta(minutes=exact_segment_minutes)
        seg_end = seg_start + timedelta(minutes=rounded_segment_minutes)
        
        classification = classify_interval(seg_start, seg_end, race_config)

        segment_rows.append({
            "team_id": team_id,
            "segment_id": seg_id,
            "algus_kp_id": start_kp_id,
            "lopp_kp_id": end_kp_id,
            "start_time": seg_start,
            "end_time": seg_end,
            "distance_m": dist_m,
            "reaalne_valgustingimus": classification,
            "exact_minutes": exact_segment_minutes,
            "minutes_total": rounded_segment_minutes,
        })

        # Lisa kontrollpunkt (va start)
        if end_kp_id > 0:
            kp_duration = float(cp_lookup[end_kp_id]["kestvus_min"])
            kp_arrival = seg_end
            kp_departure = kp_arrival + timedelta(minutes=kp_duration)
            checkpoint_rows.append({
                "team_id": team_id,
                "kp_id": end_kp_id,
                "kp_nimi": cp_lookup[end_kp_id]["nimi"],
                "arrival_time": kp_arrival,
                "departure_time": kp_departure
            })
            current_time = kp_departure
        else:
            current_time = seg_end

    return pd.DataFrame(segment_rows), pd.DataFrame(checkpoint_rows)


def simulate_all_teams(control_points: pd.DataFrame, segments: pd.DataFrame, race_config: dict, start_duration_min: int) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    start_times_df = generate_team_start_times(race_config)
    all_segment_results = []
    all_checkpoint_results = []

    for _, team_row in start_times_df.iterrows():
        team_id = int(team_row["team_id"])
        team_start = pd.to_datetime(team_row["start_time"])

        seg_res, cp_res = simulate_team_route(
            team_id=team_id,
            team_start=team_start,
            control_points=control_points,
            segments=segments,
            race_config=race_config,
            start_duration_min=start_duration_min,
        )
        seg_res = seg_res.merge(
            segments[["segment_id", "valgustingimused"]].rename(columns={"valgustingimused": "chosen_light_condition"}),
            on="segment_id",
            how="left"
        )

        all_segment_results.append(seg_res)
        all_checkpoint_results.append(cp_res)

    segment_results_df = pd.concat(all_segment_results, ignore_index=True)
    checkpoint_results_df = pd.concat(all_checkpoint_results, ignore_index=True)
    return start_times_df, segment_results_df, checkpoint_results_df


def run_full_simulation(control_points_input: pd.DataFrame, segments_input: pd.DataFrame, race_config: dict, default_speeds: dict, overrides: Dict[int, Dict[str, float]], start_mgrs: str, start_duration_min: int):
    validate_inputs(control_points_input, segments_input, race_config)

    # Lisa start kontrollpunktide hulka
    start_lat, start_lon = mgrs_to_latlon(start_mgrs)
    start_row = pd.DataFrame([{
        "kp_id": 0,
        "nimi": "Start",
        "mgrs": start_mgrs,
        "kestvus_ettevalmistus_min": 0,
        "kestvus_uleanne_min": start_duration_min,
        "kestvus_tagasiside_min": 0,
        "jarjekord": 0
    }])
    cp_input = pd.concat([start_row, control_points_input], ignore_index=True)
    cp_input = normalize_control_point_durations(cp_input)

    if race_config.get("kasuta_automaatset_paikest", False):
        race_config = compute_sun_times(cp_input, race_config)

    cp = enrich_control_points_with_coordinates(cp_input)
    seg = calculate_segment_distances(cp, segments_input)
    seg = apply_speeds(seg, default_speeds, overrides)

    start_times_df = generate_team_start_times(race_config)
    current_team_times = [
        pd.to_datetime(start_time).to_pydatetime() + timedelta(minutes=start_duration_min)
        for start_time in start_times_df["start_time"]
    ]
    cp_lookup = cp.set_index("kp_id").to_dict("index")

    chosen_speeds = []
    light_conditions = []
    exact_segment_minutes = []
    rounded_segment_minutes = []
    for _, row in seg.sort_values("segment_id").iterrows():
        dist_m = float(row["kasutatav_kaugus_m"])
        white_speed = float(row["kiirus_valges_kmh"])
        dark_speed = float(row["kiirus_pimedas_kmh"])

        white_exact_minutes = calculate_segment_minutes(dist_m, white_speed)
        white_rounded_minutes = round_up_minutes_to_next_five(white_exact_minutes)
        white_classifications = [
            classify_interval(team_time, team_time + timedelta(minutes=white_rounded_minutes), race_config)
            for team_time in current_team_times
        ]

        if all(classification == "valge" for classification in white_classifications):
            seg_condition = "valge"
            chosen_speed = white_speed
            exact_minutes = white_exact_minutes
            rounded_minutes = white_rounded_minutes
        else:
            seg_condition = "pime"
            chosen_speed = dark_speed
            exact_minutes = calculate_segment_minutes(dist_m, dark_speed)
            rounded_minutes = round_up_minutes_to_next_five(exact_minutes)

        chosen_speeds.append(chosen_speed)
        light_conditions.append(seg_condition)
        exact_segment_minutes.append(exact_minutes)
        rounded_segment_minutes.append(rounded_minutes)

        seg_duration = timedelta(minutes=rounded_minutes)
        if int(row["lopp_kp_id"]) > 0:
            kp_duration = float(cp_lookup[int(row["lopp_kp_id"])]["kestvus_min"])
            stop_duration = timedelta(minutes=kp_duration)
        else:
            stop_duration = timedelta(0)
        current_team_times = [
            team_time + seg_duration + stop_duration
            for team_time in current_team_times
        ]

    seg = seg.sort_values("segment_id").reset_index(drop=True)
    seg["chosen_speed_kmh"] = chosen_speeds
    seg["valgustingimused"] = light_conditions
    seg["exact_segment_minutes"] = exact_segment_minutes
    seg["rounded_segment_minutes"] = rounded_segment_minutes

    start_times_df, segment_results_df, checkpoint_results_df = simulate_all_teams(cp, seg, race_config, start_duration_min)

    # Lisa start koordinaadid
    results = {
        "race_config": race_config,
        "control_points": cp,
        "segments": seg,
        "start_times": start_times_df,
        "segment_results": segment_results_df,
        "checkpoint_results": checkpoint_results_df,
        "start_lat": start_lat,
        "start_lon": start_lon,
    }
    return results


# ======================================================
# SISENDI VALIDEERIMINE
# ======================================================

def validate_inputs(control_points: pd.DataFrame, segments: pd.DataFrame, race_config: dict):
    required_cp = {"kp_id", "nimi", "mgrs", "jarjekord"}
    required_seg = {"segment_id", "algus_kp_id", "lopp_kp_id", "liikumisviis"}

    if control_points.empty:
        raise ValueError("Kontrollpunktide tabel ei tohi olla tühi.")
    if segments.empty:
        raise ValueError("Lõikude tabel ei tohi olla tühi.")

    if not required_cp.issubset(control_points.columns):
        missing = required_cp - set(control_points.columns)
        raise ValueError(f"Kontrollpunktide tabelist puuduvad veerud: {missing}")

    if not ("kestvus_min" in control_points.columns or {"kestvus_ettevalmistus_min", "kestvus_uleanne_min", "kestvus_tagasiside_min"}.issubset(control_points.columns)):
        raise ValueError("Kontrollpunktide tabel peab sisaldama kas kestvus_min või kestvus_ettevalmistus_min;kestvus_uleanne_min;kestvus_tagasiside_min")

    if not required_seg.issubset(segments.columns):
        missing = required_seg - set(segments.columns)
        raise ValueError(f"Lõikude tabelist puuduvad veerud: {missing}")

    control_points = normalize_control_point_durations(control_points)

    if control_points[list(required_cp)].isna().any().any():
        raise ValueError("Kontrollpunktide kohustuslikud väljad ei tohi olla tühjad.")
    if segments[list(required_seg)].isna().any().any():
        raise ValueError("Lõikude kohustuslikud väljad ei tohi olla tühjad.")

    if control_points["kp_id"].duplicated().any():
        raise ValueError("kp_id peab olema unikaalne.")

    if control_points["jarjekord"].duplicated().any():
        raise ValueError("jarjekord peab olema unikaalne.")

    if segments["segment_id"].duplicated().any():
        raise ValueError("segment_id peab olema unikaalne.")

    if (control_points["kestvus_min"] < 0).any():
        raise ValueError("Kontrollpunkti kestvus ei tohi olla negatiivne.")
    if (control_points[["kestvus_ettevalmistus_min", "kestvus_uleanne_min", "kestvus_tagasiside_min"]].fillna(0) < 0).any().any():
        raise ValueError("Kontrollpunkti kestvuse alamajad ei tohi olla negatiivsed.")

    allowed_modes = {"tee", "varjatud"}
    if not segments["liikumisviis"].isin(allowed_modes).all():
        wrong = segments.loc[~segments["liikumisviis"].isin(allowed_modes), "liikumisviis"].tolist()
        raise ValueError(f"Lubatud liikumisviisid on ainult {allowed_modes}. Vigased väärtused: {wrong}")

    existing_kp_ids = set(control_points["kp_id"]) | {0}
    if not segments["algus_kp_id"].isin(existing_kp_ids).all():
        raise ValueError("Kõik algus_kp_id väärtused peavad viitama olemasolevale kontrollpunktile.")
    if not segments["lopp_kp_id"].isin(existing_kp_ids).all():
        raise ValueError("Kõik lopp_kp_id väärtused peavad viitama olemasolevale kontrollpunktile.")

    if race_config["voistkondade_arv"] < 1:
        raise ValueError("Võistkondade arv peab olema vähemalt 1.")
    if race_config["stardi_intervall_min"] <= 0:
        raise ValueError("Stardiintervall peab olema suurem kui 0.")
    if race_config["nulltiimi_earlier_min"] < 0:
        raise ValueError("0-tiimi eelaeg ei tohi olla negatiivne.")

    parse_datetime(race_config["esimese_voistkonna_start"])
    parse_date(race_config["voistluse_kuupaev"])
    if not race_config.get("kasuta_automaatset_paikest", False):
        parse_clock(race_config["paeva_algus"])
        parse_clock(race_config["pimeduse_algus"])


# ======================================================
# KOORMUSE ANALÜÜS
# ======================================================

def compute_checkpoint_load(checkpoint_results: pd.DataFrame) -> pd.DataFrame:
    records = []
    for kp_id, group in checkpoint_results.groupby("kp_id"):
        events = []
        kp_name = group["kp_nimi"].iloc[0]
        for _, row in group.iterrows():
            events.append((pd.to_datetime(row["arrival_time"]), +1))
            events.append((pd.to_datetime(row["departure_time"]), -1))
        events.sort(key=lambda x: (x[0], -x[1]))

        current_load = 0
        max_load = 0
        max_time = None
        for ts, delta in events:
            current_load += delta
            if current_load > max_load:
                max_load = current_load
                max_time = ts

        records.append({
            "kp_id": kp_id,
            "kp_nimi": kp_name,
            "maksimaalne_koormus": max_load,
            "maks_koormuse_hetk": max_time,
            "kokku_kulastusi": group.shape[0],
            "varaseim_saabumine": pd.to_datetime(group["arrival_time"]).min(),
            "hiliseim_lahkumine": pd.to_datetime(group["departure_time"]).max(),
        })

    return pd.DataFrame(records).sort_values("kp_id").reset_index(drop=True)


# ======================================================
# VÄLJUNDITABELID
# ======================================================

def format_output_tables(results: dict):
    cp = results["control_points"].copy()
    seg = results["segments"].copy()
    starts = results["start_times"].copy()
    seg_res = results["segment_results"].copy()
    cp_res = results["checkpoint_results"].copy()
    kp_load = compute_checkpoint_load(results["checkpoint_results"])

    if "reaalne_valgustingimus" not in seg_res.columns:
        if "light_classification" in seg_res.columns:
            seg_res["reaalne_valgustingimus"] = seg_res["light_classification"]
        elif "start_time" in seg_res.columns and "end_time" in seg_res.columns:
            race_config = results.get("race_config", {})
            try:
                seg_res["reaalne_valgustingimus"] = seg_res.apply(
                    lambda row: classify_interval(
                        pd.to_datetime(row["start_time"]),
                        pd.to_datetime(row["end_time"]),
                        race_config
                    ),
                    axis=1,
                )
            except Exception:
                seg_res["reaalne_valgustingimus"] = ""
        else:
            seg_res["reaalne_valgustingimus"] = ""

    for df in [starts, seg_res, cp_res, kp_load]:
        for col in df.columns:
            if "time" in col or "saabumine" in col or "lahkumine" in col or "hetk" in col:
                try:
                    df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    pass

    seg["sirge_kaugus_km"] = (seg["sirge_kaugus_m"] / 1000).round(2)
    seg["tee_kaugus_km"] = (seg["tee_kaugus_m"] / 1000).round(2)
    seg["kasutatav_kaugus_km"] = (seg["kasutatav_kaugus_m"] / 1000).round(2)

    seg["liikumiskiirus"] = ""

    segment_lighting = (
        seg_res.groupby("segment_id")["reaalne_valgustingimus"]
        .agg(lambda x: sorted(set(x)))
        .reset_index()
    )
    mixed_counts = (
        seg_res[seg_res["reaalne_valgustingimus"].isin(["sega", "segalõik"])]
        .groupby("segment_id")
        .size()
        .reset_index(name="mixed_team_count")
    )

    def format_lighting(row):
        lighting_types = row["reaalne_valgustingimus"]
        if "sega" in lighting_types or "segalõik" in lighting_types or len(lighting_types) > 1:
            if "sega" in lighting_types or "segalõik" in lighting_types:
                return f"sega ({int(row.get('mixed_team_count', 0))})"
            return "sega"
        if lighting_types == ["valge"]:
            return "valge"
        if lighting_types == ["pime"]:
            return "pime"
        return ", ".join(lighting_types)

    segment_lighting = segment_lighting.merge(mixed_counts, on="segment_id", how="left")
    segment_lighting["valgustingimused"] = segment_lighting.apply(format_lighting, axis=1)

    seg = seg.drop(columns=["valgustingimused"], errors="ignore")
    seg = seg.merge(segment_lighting[["segment_id", "valgustingimused"]], on="segment_id", how="left")

    def format_speed(row):
        lighting = row.get("valgustingimused")
        if lighting == "valge":
            return f"{row['kiirus_valges_kmh']:.1f}"
        return f"{row['kiirus_pimedas_kmh']:.1f}"

    seg["liikumiskiirus"] = seg.apply(format_speed, axis=1)
    if "distance_note" not in seg.columns:
        seg["distance_note"] = ""

    segment_summary = (
        seg_res.groupby("segment_id", as_index=False)
        .agg({"exact_minutes": "first", "minutes_total": "first"})
    )
    segment_summary["liikumise aeg täpne (min)"] = segment_summary["exact_minutes"].round(2)
    segment_summary["liikumise aeg ümardatud (min)"] = segment_summary["minutes_total"].round(0)
    seg = seg.merge(
        segment_summary[["segment_id", "liikumise aeg täpne (min)", "liikumise aeg ümardatud (min)"]],
        on="segment_id",
        how="left",
    )
    seg_res["distance_km"] = (seg_res["distance_m"] / 1000).round(2)
    seg_res["minutes_total"] = seg_res["minutes_total"].round(0)

    cp_sync = results["checkpoint_results"].copy()
    cp_sync["arrival_time"] = pd.to_datetime(cp_sync["arrival_time"]).dt.strftime("%H:%M")
    cp_sync["departure_time"] = pd.to_datetime(cp_sync["departure_time"]).dt.strftime("%H:%M")
    cp_sync["time"] = cp_sync["arrival_time"] + "/" + cp_sync["departure_time"]
    # Ensure all required columns are present in seg
    required_seg_cols = [
        "valgustingimused", "liikumiskiirus", "liikumise aeg täpne (min)", 
        "liikumise aeg ümardatud (min)", "distance_note"
    ]
    for col in required_seg_cols:
        if col not in seg.columns:
            seg[col] = ""

    return cp, seg, starts, seg_res, cp_res, kp_load, cp_sync, create_sync_diagram(cp_res, results["race_config"])


def create_sync_diagram(checkpoint_results: pd.DataFrame, race_config: dict):
    df = checkpoint_results.copy()
    df['Start'] = pd.to_datetime(df['arrival_time'])
    df['Finish'] = pd.to_datetime(df['departure_time'])
    df['Team'] = df['team_id'].astype(str)
    df['Checkpoint'] = df['kp_id'].astype(str)
    df['TaskLabel'] = df.apply(lambda row: f"KP {row['kp_id']}", axis=1)
    df = df.sort_values(['team_id', 'Start'])

    start_times = generate_team_start_times(race_config)
    timeline_start = pd.to_datetime(start_times["start_time"]).min().to_pydatetime() - timedelta(hours=1)
    timeline_end = df["Finish"].max().to_pydatetime() + timedelta(hours=1)

    fig = px.timeline(
        df,
        x_start="Start",
        x_end="Finish",
        y="Team",
        color="Checkpoint",
        hover_name="TaskLabel",
        title="Kontrollpunktide sünkroniseerimine",
        category_orders={"Team": sorted(df['Team'].unique(), key=lambda x: int(x))}
    )

    fig.update_yaxes(autorange="reversed", title_text="Võistkond")
    fig.update_layout(
        xaxis_title="Aeg",
        legend_title_text="Kontrollpunkt",
        plot_bgcolor='white',
        bargap=0.15,
        title_x=0.02,
    )
    fig.update_xaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgrey',
        tickformat='%d.%m<br>%H:%M',
        hoverformat='%d.%m.%Y %H:%M',
        range=[timeline_start, timeline_end],
        tick0=timeline_start,
        dtick=3600000,
        minor=dict(
            showgrid=True,
            gridwidth=0.5,
            gridcolor='rgba(200,200,200,0.4)',
            dtick=300000
        )
    )

    y1 = len(df['Team'].unique()) - 0.5

    def add_period_shape(period_start, period_end, fillcolor, label):
        x0 = max(period_start, timeline_start)
        x1 = min(period_end, timeline_end)
        if x0 >= x1:
            return
        fig.add_shape(type='rect', x0=x0, x1=x1, y0=-0.5, y1=y1, fillcolor=fillcolor, line_width=0)
        fig.add_annotation(
            x=x0 + (x1 - x0) / 2,
            y=0,
            text=label,
            showarrow=False,
            yanchor='bottom',
            font=dict(color='black', size=10),
            bgcolor='rgba(255,255,255,0.5)'
        )

    current_date = timeline_start.date() - timedelta(days=1)
    last_date = timeline_end.date() + timedelta(days=1)
    while current_date <= last_date:
        sunrise, sunset = get_sun_period_datetimes_for_date(race_config, current_date)
        next_sunrise, _ = get_sun_period_datetimes_for_date(race_config, current_date + timedelta(days=1))
        dawn_start = sunrise - timedelta(hours=1)
        dusk_end = sunset + timedelta(hours=1)
        dark_start = dusk_end
        dark_end = next_sunrise - timedelta(hours=1)

        add_period_shape(dawn_start, sunrise, 'rgba(255, 223, 186, 0.18)', 'Hommikune hämar aeg')
        add_period_shape(sunset, dusk_end, 'rgba(255, 223, 186, 0.18)', 'Hämar aeg')
        add_period_shape(dark_start, dark_end, 'rgba(200, 200, 200, 0.14)', 'Pime aeg')

        if timeline_start <= sunrise <= timeline_end:
            fig.add_vline(x=sunrise, line=dict(color='goldenrod', width=2, dash='dash'))
            fig.add_annotation(
                x=sunrise,
                y=-0.5,
                text=f"Päikesetõus<br>{sunrise.strftime('%d.%m %H:%M')}",
                showarrow=False,
                yshift=-26,
                font=dict(color='goldenrod', size=10),
            )
        if timeline_start <= sunset <= timeline_end:
            fig.add_vline(x=sunset, line=dict(color='goldenrod', width=2, dash='dash'))
            fig.add_annotation(
                x=sunset,
                y=-0.5,
                text=f"Päikeseloojang<br>{sunset.strftime('%d.%m %H:%M')}",
                showarrow=False,
                yshift=-26,
                font=dict(color='goldenrod', size=10),
            )

        current_date += timedelta(days=1)

    fig.add_trace(go.Scatter(
        x=[None], y=[None],
        mode='markers',
        marker=dict(size=10, color='rgba(255, 223, 186, 0.5)'),
        legendgroup='twilight',
        name='Hämar aeg',
        showlegend=True,
        hoverinfo='none'
    ))
    fig.add_trace(go.Scatter(
        x=[None], y=[None],
        mode='markers',
        marker=dict(size=10, color='rgba(200, 200, 200, 0.5)'),
        legendgroup='dark',
        name='Pime aeg',
        showlegend=True,
        hoverinfo='none'
    ))

    return fig


def summarize_segment_classifications(segment_results_df: pd.DataFrame) -> pd.DataFrame:
    df = segment_results_df.copy()
    if "reaalne_valgustingimus" not in df.columns:
        if "light_classification" in df.columns:
            df["reaalne_valgustingimus"] = df["light_classification"]
        else:
            df["reaalne_valgustingimus"] = ""

    summary = (
        df
        .groupby(["segment_id", "reaalne_valgustingimus"])
        .size()
        .reset_index(name="team_count")
    )
    summary["reaalne_valgustingimus"] = summary["reaalne_valgustingimus"].replace({"segalõik": "sega"})
    pivot = summary.pivot(index="segment_id", columns="reaalne_valgustingimus", values="team_count").fillna(0)
    if not pivot.empty:
        try:
            pivot = pivot.astype(int)
        except Exception:
            pass
    pivot = pivot.rename_axis(None, axis=1).reset_index()
    if "valge" not in pivot.columns:
        pivot["valge"] = 0
    if "pime" not in pivot.columns:
        pivot["pime"] = 0
    if "sega" not in pivot.columns:
        pivot["sega"] = 0
    return pivot[["segment_id", "valge", "pime", "sega"]]


# ======================================================
# KAARDIVADE
# ======================================================

def create_map(control_points: pd.DataFrame, segments: pd.DataFrame, checkpoint_results: Optional[pd.DataFrame] = None, segment_results: Optional[pd.DataFrame] = None):
    cp = control_points.sort_values("jarjekord").reset_index(drop=True)
    center_lat = cp["lat"].mean()
    center_lon = cp["lon"].mean()

    m = folium.Map(location=[center_lat, center_lon], zoom_start=12)

    if segment_results is not None and not segment_results.empty:
        if "reaalne_valgustingimus" not in segment_results.columns:
            segment_results = segment_results.copy()
            if "light_classification" in segment_results.columns:
                segment_results["reaalne_valgustingimus"] = segment_results["light_classification"].astype(str)
            else:
                segment_results["reaalne_valgustingimus"] = ""
        else:
            segment_results = segment_results.copy()

    for _, row in cp.iterrows():
        kp_id = int(row["kp_id"])
        duration_text = f"{row['kestvus_min']} min"
        if all(col in row and pd.notna(row[col]) for col in ["kestvus_ettevalmistus_min", "kestvus_uleanne_min", "kestvus_tagasiside_min"]):
            duration_text = (
                f"{row['kestvus_min']} min (Ettevalmistus: {int(row['kestvus_ettevalmistus_min'])} min, "
                f"Ülesanne: {int(row['kestvus_uleanne_min'])} min, "
                f"Tagasiside: {int(row['kestvus_tagasiside_min'])} min)"
            )
        
        tooltip_text = row["nimi"]
        
        if checkpoint_results is not None and not checkpoint_results.empty:
            cp_filter = checkpoint_results[checkpoint_results["kp_id"] == kp_id]
            if not cp_filter.empty:
                first_arrival = cp_filter["arrival_time"].min()
                last_departure = cp_filter["departure_time"].max()
                popup = (
                    f"<b>{row['nimi']}</b><br>"
                    f"KP ID: {row['kp_id']}<br>"
                    f"MGRS: {row['mgrs']}<br>"
                    f"1. võistkonna saabumine: {pd.to_datetime(first_arrival).strftime('%H:%M')}<br>"
                    f"Viimase võistkonna lõpetamine: {pd.to_datetime(last_departure).strftime('%H:%M')}<br>"
                    f"Ettevalmistus: {int(row.get('kestvus_ettevalmistus_min', 0))} min<br>"
                    f"Ülesanne: {int(row.get('kestvus_uleanne_min', 0))} min<br>"
                    f"Tagasiside: {int(row.get('kestvus_tagasiside_min', 0))} min"
                )
            else:
                popup = (
                    f"<b>{row['nimi']}</b><br>"
                    f"KP ID: {row['kp_id']}<br>"
                    f"MGRS: {row['mgrs']}<br>"
                    f"Kestvus: {duration_text}"
                )
        else:
            popup = (
                f"<b>{row['nimi']}</b><br>"
                f"KP ID: {row['kp_id']}<br>"
                f"MGRS: {row['mgrs']}<br>"
                f"Kestvus: {duration_text}"
            )

        folium.Marker(
            location=[row["lat"], row["lon"]],
            popup=popup,
            tooltip=tooltip_text,
        ).add_to(m)

    cp_lookup = cp.set_index("kp_id").to_dict("index")
    for _, seg in segments.iterrows():
        start_cp = cp_lookup[seg["algus_kp_id"]]
        end_cp = cp_lookup[seg["lopp_kp_id"]]
        if "route_coords" in seg and seg["route_coords"] is not None:
            points = [[lat, lon] for lat, lon in seg["route_coords"]]
        else:
            points = [[start_cp["lat"], start_cp["lon"]], [end_cp["lat"], end_cp["lon"]]]
        
        color = "blue" if seg["liikumisviis"] == "tee" else "red"
        
        tooltip_text = f"Lõik {seg['segment_id']} ({seg['liikumisviis']})"
        
        popup = (
            f"<b>Lõik {seg['segment_id']}</b><br>"
            f"Tüüp: {seg['liikumisviis']}<br>"
            f"Kaugus: {seg['kasutatav_kaugus_m']/1000:.2f} km<br>"
            f"Kiirus päeval: {seg['kiirus_valges_kmh']:.1f} km/h<br>"
            f"Kiirus öösel: {seg['kiirus_pimedas_kmh']:.1f} km/h"
        )
        
        if segment_results is not None and not segment_results.empty:
            seg_filter = segment_results[segment_results["segment_id"] == seg["segment_id"]].copy()
            if not seg_filter.empty:
                first_start = seg_filter["start_time"].min() if "start_time" in seg_filter.columns else None
                last_end = seg_filter["end_time"].max() if "end_time" in seg_filter.columns else None
                if "reaalne_valgustingimus" not in seg_filter.columns:
                    if "light_classification" in seg_filter.columns:
                        seg_filter["reaalne_valgustingimus"] = seg_filter["light_classification"].astype(str)
                    else:
                        seg_filter["reaalne_valgustingimus"] = ""
                try:
                    light_class_counts = (
                        seg_filter.loc[seg_filter["reaalne_valgustingimus"].astype(str).ne(""), "reaalne_valgustingimus"]
                        .value_counts()
                        .to_dict()
                    )
                except KeyError:
                    light_class_counts = {}
                if first_start is not None:
                    popup += f"<br>1. võistkonna start: {pd.to_datetime(first_start).strftime('%H:%M')}"
                if last_end is not None:
                    popup += f"<br>Viimane lõpetab: {pd.to_datetime(last_end).strftime('%H:%M')}"
                if light_class_counts:
                    light_summary = ", ".join([f"{k}: {v}" for k, v in sorted(light_class_counts.items())])
                    popup += f"<br>Valgusklassid: {light_summary}"

        folium.PolyLine(points, color=color, popup=popup, tooltip=tooltip_text).add_to(m)

    return m


# ======================================================
# EXCEL EKSPORT
# ======================================================

def format_excel_datetime(value) -> str:
    if value is None or pd.isna(value):
        return ""
    dt = pd.to_datetime(value)
    return f"{dt.day}.{dt.month}.{str(dt.year)[2:]} {dt.hour}:{dt.minute:02d}"


def format_excel_time(value) -> str:
    if value is None or pd.isna(value):
        return ""
    dt = pd.to_datetime(value)
    return f"{dt.hour:02d}:{dt.minute:02d}"


def format_duration_hms(minutes_value) -> str:
    if minutes_value is None or pd.isna(minutes_value):
        return ""
    total_seconds = int(round(float(minutes_value) * 60))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"


def team_display_name(team_id: int) -> str:
    return f"VK {team_id}"


def build_team_schedule(results: dict, team_id: int) -> List[dict]:
    cp = results["control_points"].copy().sort_values("jarjekord")
    segments = results["segments"].copy().sort_values("segment_id")
    checkpoint_results = results["checkpoint_results"].copy()
    start_times = results["start_times"].copy()

    cp_lookup = cp.set_index("kp_id").to_dict("index")
    segment_by_start = {int(row["algus_kp_id"]): row for _, row in segments.iterrows()}
    team_start = pd.to_datetime(start_times.loc[start_times["team_id"] == team_id, "start_time"].iloc[0])
    team_checkpoints = checkpoint_results[checkpoint_results["team_id"] == team_id].copy()
    checkpoint_by_id = {int(row["kp_id"]): row for _, row in team_checkpoints.iterrows()}

    rows = []
    for _, cp_row in cp.iterrows():
        kp_id = int(cp_row["kp_id"])
        if kp_id == 0:
            prep_start = team_start
            checkpoint_row = None
        else:
            checkpoint_row = checkpoint_by_id[kp_id]
            prep_start = pd.to_datetime(checkpoint_row["arrival_time"])

        prep_minutes = float(cp_row.get("kestvus_ettevalmistus_min", 0) or 0)
        task_minutes = float(cp_row.get("kestvus_uleanne_min", 0) or 0)
        task_start = prep_start + timedelta(minutes=prep_minutes)
        task_end = task_start + timedelta(minutes=task_minutes)
        departure_time = task_end if checkpoint_row is None else pd.to_datetime(checkpoint_row["departure_time"])

        next_segment = segment_by_start.get(kp_id)
        movement_minutes = None
        if next_segment is not None:
            movement_minutes = float(next_segment.get("rounded_segment_minutes", next_segment.get("minutes_total", 0)) or 0)

        rows.append({
            "kp_id": kp_id,
            "kp_label": "Start" if kp_id == 0 else str(kp_id),
            "nimi": cp_row.get("nimi", ""),
            "koordinaat": cp_row.get("mgrs", ""),
            "ettevalmistuste_algus": prep_start,
            "ulesande_algus": task_start,
            "ulesande_lopp": task_end,
            "kontrollpunkti_lopp": departure_time,
            "liikumisaeg_jargmisesse": movement_minutes,
            "markused": "",
        })
    return rows


def style_range(ws, min_row, max_row, min_col, max_col, border, fill=None, font=None, alignment=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment


def write_data_sheet(ws, results: dict, styles: dict):
    segments = results["segments"].copy().sort_values("segment_id")
    cp = results["control_points"].copy().sort_values("jarjekord")
    starts = results["start_times"].copy().sort_values("team_id")
    race_config = results["race_config"]

    border = styles["border"]
    yellow = styles["yellow"]
    bold = styles["bold"]
    center = styles["center"]

    ws["A1"] = "Liikumis ajad (h:mm:ss)"
    ws["D1"] = "Ettevalmistus aeg (h:mm:ss)"
    ws["G1"] = "Ülesande aeg (h:mm:ss)"
    ws["J1"] = "Märkused"
    ws["M1"] = "Nimed"
    for cell in ["A1", "D1", "G1", "J1", "M1"]:
        ws[cell].font = bold

    row = 2
    for _, segment in segments.iterrows():
        start_id = int(segment["algus_kp_id"])
        end_id = int(segment["lopp_kp_id"])
        start_label = "Start" if start_id == 0 else str(start_id)
        ws.cell(row=row, column=1, value=f"{start_label}->{end_id}")
        ws.cell(row=row, column=2, value=format_duration_hms(segment.get("rounded_segment_minutes", 0)))
        row += 1
    style_range(ws, 1, max(row - 1, 1), 1, 2, border)

    row = 2
    for _, cp_row in cp[cp["kp_id"] > 0].iterrows():
        ws.cell(row=row, column=4, value=int(cp_row["kp_id"]))
        ws.cell(row=row, column=5, value=format_duration_hms(cp_row.get("kestvus_ettevalmistus_min", 0)))
        row += 1
    style_range(ws, 1, max(row - 1, 1), 4, 5, border)

    row = 2
    for _, cp_row in cp[cp["kp_id"] > 0].iterrows():
        ws.cell(row=row, column=7, value=int(cp_row["kp_id"]))
        ws.cell(row=row, column=8, value=format_duration_hms(cp_row.get("kestvus_uleanne_min", 0)))
        row += 1
    style_range(ws, 1, max(row - 1, 1), 7, 8, border)

    row = 2
    for _, cp_row in cp[cp["kp_id"] > 0].iterrows():
        ws.cell(row=row, column=10, value=int(cp_row["kp_id"]))
        ws.cell(row=row, column=11, value="")
        row += 1
    style_range(ws, 1, max(row - 1, 1), 10, 11, border)

    row = 2
    for _, start_row in starts.iterrows():
        team_id = int(start_row["team_id"])
        ws.cell(row=row, column=13, value=team_display_name(team_id))
        ws.cell(row=row, column=14, value=team_display_name(team_id))
        row += 1
    style_range(ws, 1, max(row - 1, 1), 13, 14, border)

    first_team_start = pd.to_datetime(starts.loc[starts["team_id"] == 1, "start_time"].iloc[0]) if (starts["team_id"] == 1).any() else pd.to_datetime(starts["start_time"].iloc[0])
    config_rows = [
        ("Esimese VK start", format_excel_datetime(first_team_start)),
        ("Intervall", format_duration_hms(race_config["stardi_intervall_min"])),
        ("0 tiim ees", format_duration_hms(race_config["nulltiimi_earlier_min"])),
        ("Päikesetõus", race_config.get("paeva_algus", "")),
        ("Päikeseloojang", race_config.get("pimeduse_algus", "")),
    ]
    for idx, (label, value) in enumerate(config_rows, start=1):
        ws.cell(row=idx, column=16, value=label)
        ws.cell(row=idx, column=17, value=value)
    style_range(ws, 1, len(config_rows), 16, 17, border, fill=yellow, font=bold)

    for col, width in {"A": 18, "B": 14, "D": 18, "E": 14, "G": 18, "H": 14, "J": 18, "K": 24, "M": 10, "N": 24, "P": 20, "Q": 18}.items():
        ws.column_dimensions[col].width = width
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = center


def write_summary_section(ws, title: str, start_row: int, values_by_team: Dict[int, List], team_ids: List[int], kp_headers: List[str], styles: dict) -> int:
    border = styles["border"]
    yellow = styles["yellow"]
    bold = styles["bold"]
    center = styles["center"]
    gray = styles["gray"]

    max_col = 2 + len(kp_headers)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    ws.cell(row=start_row, column=1, value=title)
    style_range(ws, start_row, start_row, 1, max_col, border, fill=yellow, font=bold, alignment=center)

    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="VK")
    ws.cell(row=header_row, column=2, value="Nimi")
    for idx, header in enumerate(kp_headers, start=3):
        ws.cell(row=header_row, column=idx, value=header)
    style_range(ws, header_row, header_row, 1, max_col, border, fill=yellow, font=bold, alignment=center)

    row = header_row + 1
    for team_id in team_ids:
        ws.cell(row=row, column=1, value=team_display_name(team_id))
        ws.cell(row=row, column=2, value=team_display_name(team_id))
        for col_idx, value in enumerate(values_by_team[team_id], start=3):
            ws.cell(row=row, column=col_idx, value=value)
        fill = gray if row % 2 == 1 else None
        style_range(ws, row, row, 1, max_col, border, fill=fill, alignment=center)
        row += 1

    return row + 1


def write_summary_sheet(ws, results: dict, styles: dict):
    from openpyxl.utils import get_column_letter

    cp = results["control_points"].copy().sort_values("jarjekord")
    starts = results["start_times"].copy().sort_values("team_id")
    team_ids = [int(team_id) for team_id in starts["team_id"]]
    kp_headers = ["Start"] + [str(int(kp_id)) for kp_id in cp.loc[cp["kp_id"] > 0, "kp_id"]]

    prep_values = {}
    task_start_values = {}
    task_end_values = {}
    departure_values = {}
    for team_id in team_ids:
        schedule = build_team_schedule(results, team_id)
        prep_values[team_id] = [format_excel_datetime(row["ettevalmistuste_algus"]) for row in schedule]
        task_start_values[team_id] = [format_excel_datetime(row["ulesande_algus"]) for row in schedule]
        task_end_values[team_id] = [format_excel_datetime(row["ulesande_lopp"]) for row in schedule]
        departure_values[team_id] = [format_excel_datetime(row["kontrollpunkti_lopp"]) for row in schedule]

    row = 1
    row = write_summary_section(ws, "Ettevalmistuste algus", row, prep_values, team_ids, kp_headers, styles)
    row = write_summary_section(ws, "Ülesande algus", row, task_start_values, team_ids, kp_headers, styles)
    row = write_summary_section(ws, "Ülesande lõpp", row, task_end_values, team_ids, kp_headers, styles)
    write_summary_section(ws, "Kontrollpunkti lõpp", row, departure_values, team_ids, kp_headers, styles)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for col_idx in range(3, 3 + len(kp_headers)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def write_print_sheet(ws, results: dict, styles: dict):
    starts = results["start_times"].copy().sort_values("team_id")
    team_ids = [int(team_id) for team_id in starts["team_id"]]
    border = styles["border"]
    bold = styles["bold"]
    gray = styles["gray"]
    center = styles["center"]

    headers = ["KP", "Koordinaat", "Ettevalmistuste algus", "Ülesande algus", "Ülesande lõpp", "Liikumisaeg järgmisesse", "Märkused"]
    row = 1
    for team_id in team_ids:
        ws.cell(row=row, column=1, value=team_display_name(team_id))
        ws.cell(row=row, column=1).font = bold
        row += 1

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=header)
        style_range(ws, row, row, 1, len(headers), border, font=bold, alignment=center)
        row += 1

        for idx, item in enumerate(build_team_schedule(results, team_id)):
            values = [
                item["kp_label"],
                "" if item["kp_id"] == 0 else item["koordinaat"],
                format_excel_datetime(item["ettevalmistuste_algus"]),
                format_excel_time(item["ulesande_algus"]),
                format_excel_time(item["ulesande_lopp"]),
                format_duration_hms(item["liikumisaeg_jargmisesse"]),
                item["markused"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=value)
            fill = gray if idx % 2 == 1 else None
            style_range(ws, row, row, 1, len(headers), border, fill=fill, alignment=center)
            row += 1

        row += 3

    widths = {"A": 16, "B": 22, "C": 24, "D": 18, "E": 18, "F": 26, "G": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def export_results_to_excel(results: dict) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output = BytesIO()

    thin = Side(style="thin", color="000000")
    styles = {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "yellow": PatternFill("solid", fgColor="FFFF00"),
        "gray": PatternFill("solid", fgColor="B7B7B7"),
        "bold": Font(bold=True),
        "center": Alignment(horizontal="center", vertical="center"),
    }

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_summary = wb.create_sheet("Üld")
    ws_print = wb.create_sheet("Print")

    write_data_sheet(ws_data, results, styles)
    write_summary_sheet(ws_summary, results, styles)
    write_print_sheet(ws_print, results, styles)

    wb.save(output)

    output.seek(0)
    return output.getvalue()


def export_variant1(results: dict) -> bytes:
    """Variant 1: Iga võistkonna eraldi leht segment_results'iga"""
    from io import BytesIO
    output = BytesIO()

    seg_res = format_output_tables(results)[3]  # seg_res_out

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for team_id in sorted(seg_res["team_id"].unique()):
            team_data = seg_res[seg_res["team_id"] == team_id]
            team_data.to_excel(writer, sheet_name=f"VK_{team_id}", index=False)

    output.seek(0)
    return output.getvalue()


def export_variant2(results: dict) -> bytes:
    """Variant 2: Iga KP eraldi leht checkpoint_results'iga"""
    from io import BytesIO
    output = BytesIO()

    cp_res = format_output_tables(results)[4]  # cp_res_out

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for kp_id in sorted(cp_res["kp_id"].unique()):
            kp_data = cp_res[cp_res["kp_id"] == kp_id]
            kp_data.to_excel(writer, sheet_name=f"KP_{kp_id}", index=False)

    output.seek(0)
    return output.getvalue()


# ======================================================
# MUUTMISFUNKTSIOONID
# ======================================================

def update_segment_speed(overrides: Dict[int, Dict[str, float]], segment_id: int, valge: Optional[float] = None, pime: Optional[float] = None):
    if segment_id not in overrides:
        overrides[segment_id] = {}
    if valge is not None:
        overrides[segment_id]["valge"] = float(valge)
    if pime is not None:
        overrides[segment_id]["pime"] = float(pime)


def update_control_point_mgrs(control_points: pd.DataFrame, kp_id: int, new_mgrs: str) -> pd.DataFrame:
    cp = control_points.copy()
    idx = cp.index[cp["kp_id"] == kp_id]
    if len(idx) == 0:
        raise ValueError(f"Kontrollpunkti kp_id={kp_id} ei leitud.")
    cp.loc[idx, "mgrs"] = new_mgrs
    return cp
